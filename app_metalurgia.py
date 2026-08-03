import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import PolynomialFeatures
from sklearn.metrics import r2_score

# ── CONFIGURACIÓN ─────────────────────────────────────────────────────────
st.set_page_config(
    page_title="SSR Puna - Modelo Metalúrgico",
    page_icon="",
    layout="wide",
    initial_sidebar_state="expanded"
)

import base64
from pathlib import Path

# ── ESTILOS ───────────────────────────────────────────────────────────────
st.markdown("""
<style>
  html, body, [class*="css"] { font-size: 16px !important; }
  p, div, span, label { font-size: 16px !important; }
  .stTabs [data-baseweb="tab"] { font-size: 17px !important; font-weight: 600; }
  .stSlider label, .stSelectbox label,
  .stMultiSelect label, .stRadio label,
  .stNumberInput label, .stDateInput label {
    font-size: 17px !important; font-weight: 500 !important;
  }
  [data-testid="stSlider"] [data-baseweb="slider"] div[role="slider"] {
    width: 24px !important; height: 24px !important;
  }
  [data-testid="stSlider"] div[data-baseweb="slider"] > div:first-child {
    height: 6px !important;
  }
  [data-testid="stSlider"] [data-testid="stTickBarMin"],
  [data-testid="stSlider"] [data-testid="stTickBarMax"] { font-size: 14px !important; }
  [data-testid="stSlider"] p { font-size: 16px !important; }
  [data-testid="stMetricLabel"] { font-size: 15px !important; }
  [data-testid="stMetricValue"] { font-size: 28px !important; }
  .stDataFrame { font-size: 16px !important; }
  [data-testid="stSidebar"] { font-size: 16px !important; }
  [data-testid="stSidebar"] label { font-size: 16px !important; }
</style>
""", unsafe_allow_html=True)


def layout_claro(height=450, titulo=None):
    base = dict(
        paper_bgcolor='#ffffff',
        plot_bgcolor='#f8f9fa',
        font=dict(color='#1a1a2e', family='Barlow, sans-serif', size=14),
        xaxis=dict(gridcolor='#e0e0e0', zerolinecolor='#cccccc', tickfont=dict(size=13)),
        yaxis=dict(gridcolor='#e0e0e0', zerolinecolor='#cccccc', tickfont=dict(size=13)),
        legend=dict(bgcolor='rgba(255,255,255,0.9)', bordercolor='#cccccc',
                    borderwidth=1, font=dict(size=13)),
        height=height
    )
    if titulo:
        base['title'] = dict(text=titulo, font=dict(size=16, color='#1a1a2e'))
    return base


FONDO_CLARO = True


# ── FUNCIONES DE DATOS ────────────────────────────────────────────────────
@st.cache_data
def cargar_datos(archivo):
    try:
        if archivo.name.endswith('.csv'):
            df = pd.read_csv(archivo)
        else:
            df = pd.read_excel(archivo)
        for col in df.columns:
            if 'date' in col.lower() or 'fecha' in col.lower():
                df[col] = pd.to_datetime(df[col])
                df = df.rename(columns={col: 'fecha'})
                break
        df = df.sort_values('fecha').reset_index(drop=True)
        return df, None
    except Exception as e:
        return None, str(e)


def get_columnas_numericas(df):
    return [c for c in df.select_dtypes(include='number').columns]


def calcular_stats(df, columnas):
    stats = []
    for col in columnas:
        vals = df[col].dropna()
        stats.append({
            'Variable'       : col,
            'N'              : len(vals),
            'Mínimo'         : round(vals.min(), 3),
            'Máximo'         : round(vals.max(), 3),
            'Promedio'       : round(vals.mean(), 3),
            'Mediana'        : round(vals.median(), 3),
            'Desv. Estándar' : round(vals.std(), 3),
            'Nulos'          : df[col].isna().sum()
        })
    return pd.DataFrame(stats)


def calcular_correlaciones(df, columnas):
    return df[columnas].corr().round(4)


def exportar_coeficientes_excel(resultado, fecha_desde, fecha_hasta):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    model    = resultado['model']
    poly     = resultado['poly']
    x_cols_r = resultado['x_cols']
    y_col_r  = resultado['y_col']
    grado_r  = resultado['grado']
    r2       = resultado['r2']
    residuos = resultado['residuos']
    df_mod   = resultado['df_modelo']
    coefs     = model.coef_
    intercept = model.intercept_
    nombres   = poly.get_feature_names_out(x_cols_r)

    def limpiar_nombre(nombre, x_cols):
        res_n = nombre
        for i, col in enumerate(x_cols):
            res_n = res_n.replace(f"x{i}", col)
        return res_n

    wb = Workbook()
    COLOR_HEADER  = "1a3a5c"
    COLOR_SUBHEAD = "2d6a9f"
    COLOR_ACCENT  = "f5820a"
    COLOR_LIGHT   = "f0f4f8"
    COLOR_WHITE   = "ffffff"
    COLOR_GREEN   = "3fb950"
    COLOR_RED     = "e63946"

    def header_style(cell, bg=COLOR_HEADER, size=12, bold=True, color="ffffff"):
        cell.font      = Font(bold=bold, size=size, color=color, name='Calibri')
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def border_thin(cell):
        thin = Side(style='thin', color='cccccc')
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def data_style(cell, bold=False, bg=COLOR_WHITE, align='left'):
        cell.font      = Font(bold=bold, size=11, name='Calibri')
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=align, vertical='center')
        border_thin(cell)

    ws1 = wb.active
    ws1.title = "Modelo"
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 45
    ws1.row_dimensions[1].height = 35

    ws1.merge_cells('A1:B1')
    cell = ws1['A1']
    cell.value = "MODELO METALÚRGICO — COEFICIENTES DE AJUSTE"
    header_style(cell, bg=COLOR_HEADER, size=14)

    ws1.merge_cells('A2:B2')
    cell = ws1['A2']
    cell.value = f"Generado: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}"
    header_style(cell, bg=COLOR_SUBHEAD, size=10, bold=False)

    ws1.merge_cells('A4:B4')
    cell = ws1['A4']
    cell.value = "CONTEXTO DEL AJUSTE"
    header_style(cell, bg=COLOR_SUBHEAD, size=11)

    contexto = [
        ("Variable salida (Y)",   y_col_r),
        ("Variables entrada (X)", ", ".join(x_cols_r)),
        ("Grado del polinomio",   str(grado_r)),
        ("Período desde",         str(fecha_desde)),
        ("Período hasta",         str(fecha_hasta)),
        ("Registros utilizados",  str(len(df_mod))),
    ]
    for i, (label, valor) in enumerate(contexto, start=5):
        ws1.row_dimensions[i].height = 20
        c_label = ws1.cell(row=i, column=1, value=label)
        c_valor = ws1.cell(row=i, column=2, value=valor)
        data_style(c_label, bold=True, bg=COLOR_LIGHT if i % 2 == 0 else COLOR_WHITE)
        data_style(c_valor, bg=COLOR_LIGHT if i % 2 == 0 else COLOR_WHITE, align='right')

    fila_met = len(contexto) + 6
    ws1.merge_cells(f'A{fila_met}:B{fila_met}')
    cell = ws1[f'A{fila_met}']
    cell.value = "MÉTRICAS DEL MODELO"
    header_style(cell, bg=COLOR_SUBHEAD, size=11)

    metricas = [
        ("R² (coeficiente de determinación)", round(r2, 6)),
        ("Error estándar de residuos",         round(float(residuos.std()), 4)),
        ("Residuo medio",                      round(float(residuos.mean()), 4)),
        ("Residuo mínimo",                     round(float(residuos.min()), 4)),
        ("Residuo máximo",                     round(float(residuos.max()), 4)),
    ]
    for i, (label, valor) in enumerate(metricas, start=fila_met + 1):
        ws1.row_dimensions[i].height = 20
        c_label = ws1.cell(row=i, column=1, value=label)
        c_valor = ws1.cell(row=i, column=2, value=valor)
        data_style(c_label, bold=True, bg=COLOR_LIGHT if i % 2 == 0 else COLOR_WHITE)
        data_style(c_valor, align='right', bg=COLOR_LIGHT if i % 2 == 0 else COLOR_WHITE)
        if label.startswith("R²"):
            color_r2 = COLOR_GREEN if r2 >= 0.7 else COLOR_ACCENT if r2 >= 0.4 else COLOR_RED
            c_valor.font = Font(bold=True, size=12, color=color_r2, name='Calibri')

    fila_coef = fila_met + len(metricas) + 2
    ws1.merge_cells(f'A{fila_coef}:B{fila_coef}')
    cell = ws1[f'A{fila_coef}']
    cell.value = "COEFICIENTES"
    header_style(cell, bg=COLOR_SUBHEAD, size=11)

    fila_coef += 1
    for col_idx, texto in enumerate(['Término', 'Coeficiente'], start=1):
        c = ws1.cell(row=fila_coef, column=col_idx, value=texto)
        header_style(c, bg=COLOR_ACCENT, size=11, color=COLOR_WHITE)

    fila_coef += 1
    c1 = ws1.cell(row=fila_coef, column=1, value='Término independiente (k)')
    c2 = ws1.cell(row=fila_coef, column=2, value=round(float(intercept), 6))
    data_style(c1, bold=True, bg=COLOR_LIGHT)
    data_style(c2, align='right', bg=COLOR_LIGHT)

    for i, (nombre, coef) in enumerate(zip(nombres, coefs)):
        fila_coef += 1
        nombre_limpio = limpiar_nombre(nombre, x_cols_r)
        bg = COLOR_WHITE if i % 2 == 0 else COLOR_LIGHT
        c1 = ws1.cell(row=fila_coef, column=1, value=nombre_limpio)
        c2 = ws1.cell(row=fila_coef, column=2, value=round(float(coef), 6))
        data_style(c1, bg=bg)
        data_style(c2, align='right', bg=bg)

    fila_eq = fila_coef + 2
    ws1.merge_cells(f'A{fila_eq}:B{fila_eq}')
    cell = ws1[f'A{fila_eq}']
    cell.value = "ECUACIÓN"
    header_style(cell, bg=COLOR_SUBHEAD, size=11)

    partes = []
    for nombre, coef in zip(nombres, coefs):
        nombre_limpio = limpiar_nombre(nombre, x_cols_r)
        signo = "+" if coef >= 0 else "−"
        partes.append(f"{signo} {abs(coef):.6f}·{nombre_limpio}")
    signo_k = "+" if intercept >= 0 else "−"
    eq_str = f"{y_col_r} = {signo_k} {abs(intercept):.4f} " + " ".join(partes)

    fila_eq += 1
    ws1.merge_cells(f'A{fila_eq}:B{fila_eq}')
    ws1.row_dimensions[fila_eq].height = 30
    cell = ws1[f'A{fila_eq}']
    cell.value = eq_str
    cell.font      = Font(bold=True, size=10, name='Courier New', color=COLOR_HEADER)
    cell.fill      = PatternFill("solid", fgColor=COLOR_LIGHT)
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    border_thin(cell)

    ws2 = wb.create_sheet("Datos")
    cols_datos = x_cols_r + [y_col_r]
    for col_idx, col_name in enumerate(cols_datos, start=1):
        c = ws2.cell(row=1, column=col_idx, value=col_name)
        header_style(c, bg=COLOR_HEADER, size=11)
        ws2.column_dimensions[get_column_letter(col_idx)].width = 22
    for row_idx, row in enumerate(df_mod[cols_datos].values, start=2):
        for col_idx, val in enumerate(row, start=1):
            c = ws2.cell(row=row_idx, column=col_idx, value=round(float(val), 4))
            bg = COLOR_LIGHT if row_idx % 2 == 0 else COLOR_WHITE
            data_style(c, bg=bg, align='right')

    ws3 = wb.create_sheet("Residuos")
    for col_idx, titulo in enumerate(
        [y_col_r + ' real', y_col_r + ' predicho', 'Residuo'], start=1
    ):
        c = ws3.cell(row=1, column=col_idx, value=titulo)
        header_style(c, bg=COLOR_HEADER, size=11)
        ws3.column_dimensions[get_column_letter(col_idx)].width = 22

    y_real = df_mod[y_col_r].values
    y_pred = resultado['y_pred']
    for row_idx, (yr, yp, res) in enumerate(zip(y_real, y_pred, residuos), start=2):
        bg = COLOR_LIGHT if row_idx % 2 == 0 else COLOR_WHITE
        c1 = ws3.cell(row=row_idx, column=1, value=round(float(yr), 4))
        c2 = ws3.cell(row=row_idx, column=2, value=round(float(yp), 4))
        c3 = ws3.cell(row=row_idx, column=3, value=round(float(res), 4))
        data_style(c1, bg=bg, align='right')
        data_style(c2, bg=bg, align='right')
        data_style(c3, bg=bg, align='right')
        if abs(res) > 2 * residuos.std():
            c3.font = Font(bold=True, size=11, color=COLOR_RED, name='Calibri')

    buffer = __import__('io').BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generar_reporte_html(df_data, resultado_modelo, cols_reporte, titulo="Reporte Metalúrgico"):
    import plotly.io as pio

    def fig_a_json(fig):
        """Serializa figura eliminando el template para evitar colores internos."""
        fig.update_layout(template=None)
        return pio.to_json(fig)

    tiene_modelo  = resultado_modelo is not None
    fecha_reporte = pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')
    cols_num      = [c for c in cols_reporte if c in df_data.columns]
    stats         = calcular_stats(df_data, cols_num)
    corr_df       = calcular_correlaciones(df_data, cols_num) if len(cols_num) >= 2 else None
    graficos      = {}

    cols_plot   = cols_num[:3]
    fig_serie   = make_subplots(rows=len(cols_plot), cols=1, shared_xaxes=True,
                                subplot_titles=cols_plot, vertical_spacing=0.08)
    colores     = ['#2d6a9f', '#f5820a', '#3fb950']
    # Convertir fechas a string para evitar formato binario
    fechas_str  = df_data['fecha'].dt.strftime('%Y-%m-%d').tolist()
    for i, col in enumerate(cols_plot):
        fig_serie.add_trace(go.Scatter(
            x=fechas_str,
            y=[float(v) for v in df_data[col].fillna(0).tolist()],
            mode='lines+markers', line=dict(color=colores[i], width=1.5),
            marker=dict(size=3), name=col
        ), row=i+1, col=1)
    fig_serie.update_layout(
        height=200 * len(cols_plot),
        paper_bgcolor='#0d1117', plot_bgcolor='#0d1117',
        font=dict(color='#8b949e'), showlegend=False
    )
    fig_serie.update_xaxes(gridcolor='#21262d')
    fig_serie.update_yaxes(gridcolor='#21262d')
    graficos['serie'] = fig_a_json(fig_serie)

    if corr_df is not None:
        fig_corr = go.Figure(go.Heatmap(
            z=corr_df.values.tolist(),
            x=corr_df.columns.tolist(),
            y=corr_df.index.tolist(),
            text=corr_df.round(3).values.tolist(),
            texttemplate='%{text}',
            colorscale=[[0,'#1a3a5c'],[0.5,'#21262d'],[1,'#f5820a']], zmin=-1, zmax=1
        ))
        fig_corr.update_layout(
            height=350, paper_bgcolor='#0d1117', plot_bgcolor='#0d1117',
            font=dict(color='#8b949e'), margin=dict(l=10, r=10, t=10, b=10)
        )
        graficos['corr'] = fig_a_json(fig_corr)

    eq_str       = ""
    r2_str       = ""
    fig_mod_json = ""
    coef_html    = ""
    nombres      = []
    x_cols_r     = []

    if tiene_modelo:
        model    = resultado_modelo['model']
        poly     = resultado_modelo['poly']
        df_mod   = resultado_modelo['df_modelo']
        x_cols_r = resultado_modelo['x_cols']
        y_col_r  = resultado_modelo['y_col']
        grado_r  = resultado_modelo['grado']
        r2       = resultado_modelo['r2']
        r2_str   = f"{r2:.4f}"
        coefs     = model.coef_
        intercept = model.intercept_
        nombres   = poly.get_feature_names_out(x_cols_r)

        partes = [f"{intercept:.4f}"]
        for nombre, coef in zip(nombres, coefs):
            signo = "+" if coef >= 0 else "-"
            partes.append(f"{signo} {abs(coef):.6f}·{nombre}")
        eq_str = f"{y_col_r} = " + " ".join(partes)

        if len(x_cols_r) == 1:
            x_vals_l  = df_mod[x_cols_r[0]].values.tolist()
            y_vals_l  = df_mod[y_col_r].values.tolist()
            x_curve   = np.linspace(min(x_vals_l), max(x_vals_l), 300).reshape(-1, 1)
            y_curve_l = model.predict(poly.transform(x_curve)).tolist()
            x_curve_l = x_curve.flatten().tolist()
            fig_mod = go.Figure()
            fig_mod.add_trace(go.Scatter(
                x=x_vals_l, y=y_vals_l, mode='markers', name='Datos',
                marker=dict(color='#1a3a5c', size=8, opacity=0.8,
                            line=dict(color='#2d6a9f', width=1)),
            ))
            fig_mod.add_trace(go.Scatter(
                x=x_curve_l, y=y_curve_l, mode='lines',
                name=f'Ajuste grado {grado_r}', line=dict(color='#e63946', width=2.5)
            ))
            fig_mod.update_layout(
                xaxis_title=x_cols_r[0], yaxis_title=y_col_r,
                paper_bgcolor='#ffffff', plot_bgcolor='#f8f9fa',
                font=dict(color='#1a1a2e', size=13),
                xaxis=dict(gridcolor='#e0e0e0'), yaxis=dict(gridcolor='#e0e0e0'),
                height=450
            )
        else:
            y_vals_l   = df_mod[y_col_r].values.tolist()
            y_pred_l   = resultado_modelo['y_pred'].tolist()
            lim        = [float(min(y_vals_l)), float(max(y_vals_l))]
            fig_mod    = go.Figure()
            fig_mod.add_trace(go.Scatter(
                x=y_vals_l, y=y_pred_l, mode='markers', name='Datos',
                marker=dict(color='#2d6a9f', size=7, opacity=0.8),
            ))
            fig_mod.add_trace(go.Scatter(
                x=lim, y=lim, mode='lines', name='Ideal',
                line=dict(color='#e63946', width=1.5, dash='dash')
            ))
            fig_mod.update_layout(
                xaxis_title=f'{y_col_r} real', yaxis_title=f'{y_col_r} predicho',
                title=f'Real vs Predicho — R² = {r2:.4f} — Variables: {", ".join(x_cols_r)}',
                paper_bgcolor='#ffffff', plot_bgcolor='#f8f9fa',
                font=dict(color='#1a1a2e', size=13),
                xaxis=dict(gridcolor='#e0e0e0'), yaxis=dict(gridcolor='#e0e0e0'),
                height=450
            )
        fig_mod_json = fig_a_json(fig_mod)

        coef_df = pd.DataFrame({
            'Término'    : ['Término independiente'] + list(nombres),
            'Coeficiente': [round(intercept, 6)] + [round(c, 6) for c in coefs]
        })
        coef_html = coef_df.to_html(index=False, border=0, classes='stats-table')

    stats_html = stats.to_html(index=False, border=0, classes='stats-table')

    # ── Script JS — cada figura se asigna a variable para evitar duplicar JSON ──
    script_lines = ["  const config = {responsive:true, displayModeBar:false};"]
    if tiene_modelo and fig_mod_json:
        script_lines.append(f"  var gmod = {fig_mod_json}; Plotly.newPlot('chart-modelo', gmod.data, gmod.layout, config);")
    script_lines.append(f"  var gser = {graficos['serie']}; Plotly.newPlot('chart-serie', gser.data, gser.layout, config);")
    if "corr" in graficos:
        script_lines.append(f"  var gcor = {graficos['corr']}; Plotly.newPlot('chart-corr', gcor.data, gcor.layout, config);")
    script_js = "\n".join(script_lines)

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>{titulo}</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/plotly.js/2.26.0/plotly.min.js"></script>
<link href="https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@600;700&family=Barlow:wght@300;400;500&family=JetBrains+Mono:wght@400;600&display=swap" rel="stylesheet">
<style>
  :root {{--bg:#0d1117;--surface:#161b22;--surface2:#1c2330;--border:#30363d;--accent:#f5820a;--text:#e6edf3;--text-muted:#8b949e;--green:#3fb950;--red:#e63946;}}
  *{{box-sizing:border-box;margin:0;padding:0;}}
  body{{background:var(--bg);color:var(--text);font-family:'Barlow',sans-serif;font-weight:300;padding:0;}}
  header{{background:var(--surface);border-bottom:1px solid var(--border);padding:20px 40px;display:flex;align-items:center;gap:16px;}}
  .h-icon{{width:38px;height:38px;background:var(--accent);border-radius:6px;display:flex;align-items:center;justify-content:center;font-size:20px;}}
  .h-title{{font-family:'Barlow Condensed',sans-serif;font-weight:700;font-size:22px;}}
  .h-sub{{font-size:11px;color:var(--text-muted);letter-spacing:1px;text-transform:uppercase;margin-top:2px;}}
  .h-date{{margin-left:auto;font-family:'JetBrains Mono',monospace;font-size:11px;color:var(--text-muted);background:var(--surface2);border:1px solid var(--border);border-radius:4px;padding:4px 10px;}}
  .content{{padding:36px 40px;max-width:1400px;margin:0 auto;}}
  .section-title{{font-family:'Barlow Condensed',sans-serif;font-size:13px;font-weight:600;letter-spacing:2px;text-transform:uppercase;color:var(--text-muted);margin:36px 0 20px;display:flex;align-items:center;gap:12px;}}
  .section-title::before,.section-title::after{{content:'';flex:1;height:1px;background:var(--border);}}
  .section-title::before{{flex:0 0 24px;}}
  .kpi-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:24px;}}
  .kpi-card{{background:var(--surface);border:1px solid var(--border);border-radius:6px;padding:16px;}}
  .kpi-label{{font-size:10px;color:var(--text-muted);text-transform:uppercase;letter-spacing:1px;margin-bottom:4px;}}
  .kpi-value{{font-family:'JetBrains Mono',monospace;font-size:22px;font-weight:600;color:var(--text);}}
  .kpi-card.green .kpi-value{{color:var(--green);}}
  .eq-box{{background:var(--surface);border:1px solid var(--border);border-radius:6px;padding:18px;margin-bottom:24px;}}
  .eq-code{{font-family:'JetBrains Mono',monospace;font-size:13px;color:var(--text);word-break:break-all;line-height:1.8;}}
  .chart-card{{background:var(--surface);border:1px solid var(--border);border-radius:8px;overflow:hidden;margin-bottom:20px;}}
  .chart-title{{font-family:'Barlow Condensed',sans-serif;font-size:12px;font-weight:600;letter-spacing:1.5px;text-transform:uppercase;color:var(--text-muted);padding:14px 18px 10px;border-bottom:1px solid var(--border);}}
  .stats-table{{width:100%;border-collapse:collapse;font-size:12px;}}
  .stats-table th{{font-family:'Barlow Condensed',sans-serif;font-size:11px;font-weight:600;letter-spacing:1px;text-transform:uppercase;color:var(--text-muted);padding:10px 18px;text-align:left;background:var(--surface2);border-bottom:1px solid var(--border);}}
  .stats-table td{{font-family:'JetBrains Mono',monospace;color:var(--text);padding:10px 18px;border-bottom:1px solid rgba(48,54,61,0.5);}}
  .stats-table tr:last-child td{{border-bottom:none;}}
  .stats-table tr:hover td{{background:var(--surface2);}}
  @media print{{header{{-webkit-print-color-adjust:exact;print-color-adjust:exact;}}.chart-card{{page-break-inside:avoid;}}}}
</style>
</head>
<body>
<header>
  <div class="h-icon">⚗</div>
  <div>
    <div class="h-title">{titulo}</div>
    <div class="h-sub">Análisis metalúrgico · {len(df_data)} registros</div>
  </div>
  <div class="h-date">Generado: {fecha_reporte}</div>
</header>
<div class="content">
  <div class="section-title">Resumen</div>
  <div class="kpi-grid">
    <div class="kpi-card"><div class="kpi-label">Registros</div><div class="kpi-value">{len(df_data)}</div></div>
    <div class="kpi-card"><div class="kpi-label">Desde</div><div class="kpi-value" style="font-size:15px">{df_data['fecha'].min().strftime('%Y-%m-%d')}</div></div>
    <div class="kpi-card"><div class="kpi-label">Hasta</div><div class="kpi-value" style="font-size:15px">{df_data['fecha'].max().strftime('%Y-%m-%d')}</div></div>
    <div class="kpi-card {'green' if tiene_modelo else ''}"><div class="kpi-label">R² modelo</div><div class="kpi-value">{r2_str if tiene_modelo else '—'}</div></div>
  </div>
  {html_section_modelo}
  <div class="section-title">Serie temporal</div>
  <div class="chart-card"><div class="chart-title">Variables vs Fecha</div><div id="chart-serie"></div></div>
  {'<div class="section-title">Correlaciones</div><div class="chart-card"><div class="chart-title">Matriz de correlación</div><div id="chart-corr"></div></div>' if corr_df is not None else ''}
  <div class="section-title">Estadísticas descriptivas</div>
  <div class="chart-card"><div class="chart-title">Resumen estadístico</div><div style="padding:8px 0">{stats_html}</div></div>
</div>
<script>
{script_js}
</script>
</body>
</html>"""
    return html

    tiene_modelo  = resultado_modelo is not None
    fecha_reporte = pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')
    cols_num      = [c for c in cols_reporte if c in df_data.columns]
    stats         = calcular_stats(df_data, cols_num)
    corr_df       = calcular_correlaciones(df_data, cols_num) if len(cols_num) >= 2 else None

    # ── Serie temporal ────────────────────────────────────────────────────
    cols_plot = cols_num[:3]
    fig_serie = make_subplots(rows=len(cols_plot), cols=1, shared_xaxes=True,
                               subplot_titles=cols_plot, vertical_spacing=0.08)
    colores = ['#2d6a9f', '#f5820a', '#3fb950']
    for i, col in enumerate(cols_plot):
        fig_serie.add_trace(go.Scatter(
            x=df_data['fecha'], y=df_data[col],
            mode='lines+markers', line=dict(color=colores[i], width=1.5),
            marker=dict(size=3), name=col
        ), row=i+1, col=1)
    fig_serie.update_layout(
        height=200 * len(cols_plot),
        paper_bgcolor='#0d1117', plot_bgcolor='#0d1117',
        font=dict(color='#8b949e'), showlegend=False,
        template=None
    )
    fig_serie.update_xaxes(gridcolor='#21262d')
    fig_serie.update_yaxes(gridcolor='#21262d')
    div_serie = fig_a_div(fig_serie, 'chart-serie')

    # ── Correlaciones ─────────────────────────────────────────────────────
    div_corr = ''
    html_section_corr = ''
    if corr_df is not None:
        fig_corr = go.Figure(go.Heatmap(
            z=corr_df.values.tolist(),
            x=corr_df.columns.tolist(),
            y=corr_df.index.tolist(),
            text=corr_df.round(3).values.tolist(),
            texttemplate='%{text}',
            colorscale=[[0,'#1a3a5c'],[0.5,'#21262d'],[1,'#f5820a']],
            zmin=-1, zmax=1
        ))
        fig_corr.update_layout(
            height=350, paper_bgcolor='#0d1117', plot_bgcolor='#0d1117',
            font=dict(color='#8b949e'), margin=dict(l=10, r=10, t=10, b=10),
            template=None
        )
        div_corr = fig_a_div(fig_corr, 'chart-corr')
        html_section_corr = f'''
  <div class="section-title">Correlaciones</div>
  <div class="chart-card">
    <div class="chart-title">Matriz de correlación</div>
    {div_corr}
  </div>'''

    # ── Modelo ────────────────────────────────────────────────────────────
    eq_str       = ""
    r2_str       = ""
    div_modelo   = ""
    coef_html    = ""
    nombres      = []
    x_cols_r     = []
    html_section_modelo = ''

    if tiene_modelo:
        model    = resultado_modelo['model']
        poly     = resultado_modelo['poly']
        df_mod   = resultado_modelo['df_modelo']
        x_cols_r = resultado_modelo['x_cols']
        y_col_r  = resultado_modelo['y_col']
        grado_r  = resultado_modelo['grado']
        r2       = resultado_modelo['r2']
        r2_str   = f"{r2:.4f}"
        coefs     = model.coef_
        intercept = model.intercept_
        nombres   = poly.get_feature_names_out(x_cols_r)

        partes = [f"{intercept:.4f}"]
        for nombre, coef in zip(nombres, coefs):
            signo = "+" if coef >= 0 else "-"
            partes.append(f"{signo} {abs(coef):.6f}·{nombre}")
        eq_str = f"{y_col_r} = " + " ".join(partes)

        if len(x_cols_r) == 1:
            x_vals  = df_mod[x_cols_r[0]].values.tolist()
            y_vals  = df_mod[y_col_r].values.tolist()
            x_curve = np.linspace(min(x_vals), max(x_vals), 300).reshape(-1, 1)
            y_curve = model.predict(poly.transform(x_curve)).tolist()
            x_curve_list = x_curve.flatten().tolist()

            fig_mod = go.Figure()
            fig_mod.add_trace(go.Scatter(
                x=x_vals, y=y_vals, mode='markers', name='Datos',
                marker=dict(color='#1a3a5c', size=8, opacity=0.8,
                            line=dict(color='#2d6a9f', width=1)),
            ))
            fig_mod.add_trace(go.Scatter(
                x=x_curve_list, y=y_curve, mode='lines',
                name=f'Ajuste grado {grado_r}',
                line=dict(color='#e63946', width=2.5)
            ))
        else:
            y_vals_real = df_mod[y_col_r].values.tolist()
            y_pred_r    = resultado_modelo['y_pred'].tolist()
            lim = [float(min(y_vals_real)), float(max(y_vals_real))]

            fig_mod = go.Figure()
            fig_mod.add_trace(go.Scatter(
                x=y_vals_real, y=y_pred_r, mode='markers', name='Datos',
                marker=dict(color='#2d6a9f', size=7, opacity=0.8),
            ))
            fig_mod.add_trace(go.Scatter(
                x=lim, y=lim, mode='lines', name='Ideal',
                line=dict(color='#e63946', width=1.5, dash='dash')
            ))

        fig_mod.update_layout(
            xaxis_title=x_cols_r[0] if len(x_cols_r) == 1 else f'{y_col_r} real',
            yaxis_title=y_col_r if len(x_cols_r) == 1 else f'{y_col_r} predicho',
            paper_bgcolor='#ffffff', plot_bgcolor='#f8f9fa',
            font=dict(color='#1a1a2e', size=13),
            xaxis=dict(gridcolor='#e0e0e0'), yaxis=dict(gridcolor='#e0e0e0'),
            height=450, template=None
        )
        div_modelo = fig_a_div(fig_mod, 'chart-modelo')

        coef_df = pd.DataFrame({
            'Término'    : ['Término independiente'] + list(nombres),
            'Coeficiente': [round(intercept, 6)] + [round(c, 6) for c in coefs]
        })
        coef_html = coef_df.to_html(index=False, border=0, classes='stats-table')

        html_section_modelo = f'''
  <div class="section-title">Modelo ajustado</div>
  <div class="eq-box">
    <div class="kpi-label" style="margin-bottom:8px">Ecuación</div>
    <div class="eq-code">{eq_str}</div>
  </div>
  <div class="chart-card">
    <div class="chart-title">Ajuste — {x_cols_r[0] if len(x_cols_r)==1 else "Real vs Predicho"}</div>
    {div_modelo}
  </div>
  <div class="chart-card" style="margin-bottom:20px">
    <div class="chart-title">Coeficientes</div>
    <div style="padding:8px 0">{coef_html}</div>
  </div>'''

    stats_html = stats.to_html(index=False, border=0, classes='stats-table')

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>{titulo}</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/plotly.js/2.26.0/plotly.min.js"></script>
<link href="https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@600;700&family=Barlow:wght@300;400;500&family=JetBrains+Mono:wght@400;600&display=swap" rel="stylesheet">
<style>
  :root {{--bg:#0d1117;--surface:#161b22;--surface2:#1c2330;--border:#30363d;--accent:#f5820a;--text:#e6edf3;--text-muted:#8b949e;--green:#3fb950;--red:#e63946;}}
  *{{box-sizing:border-box;margin:0;padding:0;}}
  body{{background:var(--bg);color:var(--text);font-family:'Barlow',sans-serif;font-weight:300;padding:0;}}
  header{{background:var(--surface);border-bottom:1px solid var(--border);padding:20px 40px;display:flex;align-items:center;gap:16px;}}
  .h-icon{{width:38px;height:38px;background:var(--accent);border-radius:6px;display:flex;align-items:center;justify-content:center;font-size:20px;}}
  .h-title{{font-family:'Barlow Condensed',sans-serif;font-weight:700;font-size:22px;}}
  .h-sub{{font-size:11px;color:var(--text-muted);letter-spacing:1px;text-transform:uppercase;margin-top:2px;}}
  .h-date{{margin-left:auto;font-family:'JetBrains Mono',monospace;font-size:11px;color:var(--text-muted);background:var(--surface2);border:1px solid var(--border);border-radius:4px;padding:4px 10px;}}
  .content{{padding:36px 40px;max-width:1400px;margin:0 auto;}}
  .section-title{{font-family:'Barlow Condensed',sans-serif;font-size:13px;font-weight:600;letter-spacing:2px;text-transform:uppercase;color:var(--text-muted);margin:36px 0 20px;display:flex;align-items:center;gap:12px;}}
  .section-title::before,.section-title::after{{content:'';flex:1;height:1px;background:var(--border);}}
  .section-title::before{{flex:0 0 24px;}}
  .kpi-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:24px;}}
  .kpi-card{{background:var(--surface);border:1px solid var(--border);border-radius:6px;padding:16px;}}
  .kpi-label{{font-size:10px;color:var(--text-muted);text-transform:uppercase;letter-spacing:1px;margin-bottom:4px;}}
  .kpi-value{{font-family:'JetBrains Mono',monospace;font-size:22px;font-weight:600;color:var(--text);}}
  .kpi-card.green .kpi-value{{color:var(--green);}}
  .eq-box{{background:var(--surface);border:1px solid var(--border);border-radius:6px;padding:18px;margin-bottom:24px;}}
  .eq-code{{font-family:'JetBrains Mono',monospace;font-size:13px;color:var(--text);word-break:break-all;line-height:1.8;}}
  .chart-card{{background:var(--surface);border:1px solid var(--border);border-radius:8px;overflow:hidden;margin-bottom:20px;}}
  .chart-title{{font-family:'Barlow Condensed',sans-serif;font-size:12px;font-weight:600;letter-spacing:1.5px;text-transform:uppercase;color:var(--text-muted);padding:14px 18px 10px;border-bottom:1px solid var(--border);}}
  .stats-table{{width:100%;border-collapse:collapse;font-size:12px;}}
  .stats-table th{{font-family:'Barlow Condensed',sans-serif;font-size:11px;font-weight:600;letter-spacing:1px;text-transform:uppercase;color:var(--text-muted);padding:10px 18px;text-align:left;background:var(--surface2);border-bottom:1px solid var(--border);}}
  .stats-table td{{font-family:'JetBrains Mono',monospace;color:var(--text);padding:10px 18px;border-bottom:1px solid rgba(48,54,61,0.5);}}
  .stats-table tr:last-child td{{border-bottom:none;}}
  .stats-table tr:hover td{{background:var(--surface2);}}
  @media print{{header{{-webkit-print-color-adjust:exact;print-color-adjust:exact;}}.chart-card{{page-break-inside:avoid;}}}}
</style>
</head>
<body>
<header>
  <div class="h-icon">⚗</div>
  <div>
    <div class="h-title">{titulo}</div>
    <div class="h-sub">Análisis metalúrgico · {len(df_data)} registros</div>
  </div>
  <div class="h-date">Generado: {fecha_reporte}</div>
</header>
<div class="content">
  <div class="section-title">Resumen</div>
  <div class="kpi-grid">
    <div class="kpi-card"><div class="kpi-label">Registros</div><div class="kpi-value">{len(df_data)}</div></div>
    <div class="kpi-card"><div class="kpi-label">Desde</div><div class="kpi-value" style="font-size:15px">{df_data['fecha'].min().strftime('%Y-%m-%d')}</div></div>
    <div class="kpi-card"><div class="kpi-label">Hasta</div><div class="kpi-value" style="font-size:15px">{df_data['fecha'].max().strftime('%Y-%m-%d')}</div></div>
    <div class="kpi-card {'green' if tiene_modelo else ''}"><div class="kpi-label">R² modelo</div><div class="kpi-value">{r2_str if tiene_modelo else '—'}</div></div>
  </div>
  {html_section_modelo}
  <div class="section-title">Serie temporal</div>
  <div class="chart-card">
    <div class="chart-title">Variables vs Fecha</div>
    {div_serie}
  </div>
  {html_section_corr}
  <div class="section-title">Estadísticas descriptivas</div>
  <div class="chart-card">
    <div class="chart-title">Resumen estadístico</div>
    <div style="padding:8px 0">{stats_html}</div>
  </div>
</div>
</body>
</html>"""
    return html


# ── SIDEBAR ───────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## Modelo Metalúrgico")
    st.markdown("---")

    # ── FUSIÓN DE ARCHIVOS (opcional) ─────────────────────────────────────
    with st.expander("🔗 Fusionar archivos Excel", expanded=False):
        st.caption("Opcional — uní hasta dos archivos por columna de fecha.")

        def leer_crudo(f):
            if f.name.endswith('.csv'):
                return pd.read_csv(f)
            return pd.read_excel(f)

        def detectar_col_fecha(df):
            for col in df.columns:
                if 'date' in col.lower() or 'fecha' in col.lower():
                    df[col] = pd.to_datetime(df[col])
                    return df.rename(columns={col: 'fecha'})
            raise ValueError(f"No se encontró columna de fecha en el archivo")

        arch1 = st.file_uploader(
            "Archivo 1 (base)",
            type=['xlsx', 'xls', 'csv'],
            key="fusion_arch1"
        )
        arch2 = st.file_uploader(
            "Archivo 2 (columnas adicionales)",
            type=['xlsx', 'xls', 'csv'],
            key="fusion_arch2"
        )

        if arch1:
            try:
                df1 = detectar_col_fecha(leer_crudo(arch1))

                if arch2:
                    df2 = detectar_col_fecha(leer_crudo(arch2))

                    # Detectar columnas duplicadas
                    cols1 = set(df1.columns) - {'fecha'}
                    cols2 = set(df2.columns) - {'fecha'}
                    duplicadas = cols1 & cols2
                    if duplicadas:
                        st.warning(
                            f"Columnas duplicadas — se agregarán sufijos _1 y _2: "
                            f"{', '.join(sorted(duplicadas))}"
                        )

                    # Outer join por fecha
                    df_fusionado = pd.merge(
                        df1, df2, on='fecha', how='outer',
                        suffixes=('_1', '_2')
                    ).sort_values('fecha').reset_index(drop=True)

                    # Resumen
                    ca, cb = st.columns(2)
                    ca.metric("Archivo 1", f"{len(df1)} filas")
                    cb.metric("Archivo 2", f"{len(df2)} filas")
                    st.metric("Resultado", f"{len(df_fusionado)} filas  ·  {len(df_fusionado.columns)-1} columnas")

                    # Fechas sin match
                    fechas_solo1 = set(df1['fecha']) - set(df2['fecha'])
                    fechas_solo2 = set(df2['fecha']) - set(df1['fecha'])
                    if fechas_solo1:
                        st.caption(f"⚠ {len(fechas_solo1)} fecha(s) solo en archivo 1 → columnas del 2 vacías")
                    if fechas_solo2:
                        st.caption(f"⚠ {len(fechas_solo2)} fecha(s) solo en archivo 2 → columnas del 1 vacías")

                    # Vista previa
                    with st.expander("Vista previa (5 filas)"):
                        st.dataframe(df_fusionado.head(5), use_container_width=True)

                    # Descarga
                    import io as _io
                    buf = _io.BytesIO()
                    df_fusionado.assign(
                        fecha=df_fusionado['fecha'].dt.strftime('%Y-%m-%d')
                    ).to_excel(buf, index=False)
                    buf.seek(0)
                    st.download_button(
                        "⬇ Descargar fusionado (.xlsx)",
                        data=buf,
                        file_name=f"fusionado_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="dl_fusionado"
                    )

                    # Cargar en la app
                    if st.button(
                        "✅ Usar este archivo en la app",
                        use_container_width=True,
                        type="primary",
                        key="btn_usar_fusionado"
                    ):
                        st.session_state['archivo_fusionado'] = df_fusionado
                        st.success("Listo — continuá desde Carga de Datos.")

                else:
                    st.info("Subí un segundo archivo para fusionar, o cargá el archivo 1 directamente abajo.")

            except Exception as e:
                st.error(f"Error al fusionar: {e}")

    st.markdown("---")

    # ── CARGA DE DATOS ────────────────────────────────────────────────────
    st.markdown("### Carga de Datos")

    # Si hay archivo fusionado en session_state, lo usa directamente
    if 'archivo_fusionado' in st.session_state:
        df_raw = st.session_state['archivo_fusionado'].copy()
        st.success(f"✓ Usando archivo fusionado — {len(df_raw)} registros")
        if st.button("✕ Quitar fusionado", key="btn_quitar_fusionado"):
            del st.session_state['archivo_fusionado']
            st.rerun()
        archivo_cargado = True
    else:
        archivo = st.file_uploader(
            "Subir Excel o CSV", type=['xlsx', 'xls', 'csv'],
            help="Una fila por día, con columna de fecha"
        )
        if archivo:
            df_raw, error = cargar_datos(archivo)
            if error:
                st.error(f"Error al cargar: {error}")
                st.stop()
            st.success(f"✓ {len(df_raw)} registros cargados")
            archivo_cargado = True
        else:
            archivo_cargado = False
            st.info("Subí un archivo para comenzar")

    if archivo_cargado:
        st.markdown("### Rango de fechas")
        fecha_min = df_raw['fecha'].min().date()
        fecha_max = df_raw['fecha'].max().date()

        col1, col2 = st.columns(2)
        with col1:
            desde = st.date_input("Desde", fecha_min, min_value=fecha_min, max_value=fecha_max)
        with col2:
            hasta = st.date_input("Hasta", fecha_max, min_value=fecha_min, max_value=fecha_max)

        mask = (df_raw['fecha'].dt.date >= desde) & (df_raw['fecha'].dt.date <= hasta)
        df   = df_raw[mask].reset_index(drop=True)
        st.caption(f"{len(df)} registros en el rango seleccionado")

        st.markdown("### ✂️ Corte train / nuevo")
        usar_corte = st.toggle(
            "Activar corte de período", value=False,
            help="Divide los datos en históricos (train) y nuevos para comparar contra el modelo"
        )

        fecha_corte = None
        if usar_corte:
            fecha_corte = st.date_input(
                "Fecha de corte",
                value=df_raw['fecha'].max().date() - pd.Timedelta(days=30),
                min_value=fecha_min, max_value=fecha_max,
                help="Datos hasta esta fecha = train (azul). Datos posteriores = nuevos (naranja)"
            )
            df_train_global = df[df['fecha'].dt.date <= fecha_corte].reset_index(drop=True)
            df_new_global   = df[df['fecha'].dt.date >  fecha_corte].reset_index(drop=True)
            st.caption(f"Train: {len(df_train_global)} registros  |  Nuevos: {len(df_new_global)} registros")
        else:
            df_train_global = df.copy()
            df_new_global   = pd.DataFrame(columns=df.columns)

        st.markdown("### Variables a explorar")
        cols_numericas = get_columnas_numericas(df)
        cols_seleccionadas = st.multiselect(
            "Seleccionar variables", cols_numericas,
            default=cols_numericas[:4] if len(cols_numericas) >= 4 else cols_numericas
        )

# ── PANTALLA DE BIENVENIDA ────────────────────────────────────────────────
if not archivo_cargado:
    st.markdown("""
    <div style="
        display: flex;
        flex-direction: column;
        align-items: center;
        justify-content: center;
        min-height: 60vh;
        text-align: center;
        padding: 40px;
    ">
        <div style="font-size: 64px; margin-bottom: 24px;">⚗</div>
        <h1 style="
            font-family: 'Calibri', sans-serif;
            font-size: 32px;
            font-weight: 700;
            color: #1a3a5c;
            margin-bottom: 12px;
        ">Herramienta de Modelado Metalúrgico</h1>
        <p style="
            font-size: 18px;
            color: #8b949e;
            margin-bottom: 40px;
            max-width: 520px;
        ">Cargá un archivo Excel o CSV desde el panel izquierdo para comenzar el análisis.</p>
        <div style="
            display: grid;
            grid-template-columns: repeat(3, 1fr);
            gap: 20px;
            max-width: 700px;
            width: 100%;
        ">
            <div style="background:#f0f4f8; border-radius:10px; padding:24px 16px;">
                <div style="font-size:28px; margin-bottom:8px;">📊</div>
                <div style="font-weight:600; color:#1a3a5c; font-size:15px; margin-bottom:6px;">Exploración</div>
                <div style="color:#8b949e; font-size:13px;">Visualizá tendencias, correlaciones y estadísticas</div>
            </div>
            <div style="background:#f0f4f8; border-radius:10px; padding:24px 16px;">
                <div style="font-size:28px; margin-bottom:8px;">📈</div>
                <div style="font-weight:600; color:#1a3a5c; font-size:15px; margin-bottom:6px;">Ajuste automático</div>
                <div style="color:#8b949e; font-size:13px;">Calculá la ecuación que mejor representa tus datos</div>
            </div>
            <div style="background:#f0f4f8; border-radius:10px; padding:24px 16px;">
                <div style="font-size:28px; margin-bottom:8px;">✏️</div>
                <div style="font-weight:600; color:#1a3a5c; font-size:15px; margin-bottom:6px;">Modelo manual</div>
                <div style="color:#8b949e; font-size:13px;">Ingresá coeficientes conocidos y simulá resultados</div>
            </div>
        </div>
        <p style="margin-top:40px; color:#cccccc; font-size:13px;">
            SSR Puna — Área Metalurgia · Versión 1.0
        </p>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# ── TABS PRINCIPALES ──────────────────────────────────────────────────────
if not cols_seleccionadas:
    st.warning("Seleccioná al menos una variable en el sidebar")
    st.stop()

# ── df_filtrado disponible para todos los tabs ────────────────────────────
df_filtrado = df.copy()

# Limpiar modelo solo si cambia el rango de fechas
clave_rango = f"{desde}_{hasta}"
if 'modelo_resultado' in st.session_state:
    if st.session_state.get('_ultimo_rango') != clave_rango:
        del st.session_state['modelo_resultado']
st.session_state['_ultimo_rango'] = clave_rango

tab1, tab2, tab3 = st.tabs([
    "> Exploración de datos",
    "> Ajuste automático",
    "> Modelo manual"
])

# ── ECUACIONES DE LABORATORIO ─────────────────────────────────────────────
ECUACIONES_LAB = {
    "Ag Recovery": {
        "k" : 0.9552,
        "c1": -0.1368,
        "c2": -0.2244,
        "c3": 0.062,
        "x_col": "lgo"
    },
    "Pb Recovery": {
        "k" : 0.0, "c1": 0.0, "c2": 0.0, "c3": 0.0, "x_col": ""
    },
}


def predict_lab(x, k, c1, c2, c3):
    x_prop = x / 100.0
    return k + c1*x_prop + c2*x_prop**2 + c3*x_prop**3


# ═════════════════════════════════════════════════════════════════════════
# TAB 1 — EXPLORACIÓN DE DATOS
# ═════════════════════════════════════════════════════════════════════════
with tab1:

    st.markdown("### Tabla de datos")

    with st.expander("Filtros adicionales por columna"):
        for col in cols_seleccionadas:
            v_min = float(df[col].min())
            v_max = float(df[col].max())
            if v_min < v_max:
                rango = st.slider(
                    col, min_value=v_min, max_value=v_max,
                    value=(v_min, v_max), key=f"filtro_{col}"
                )
                df_filtrado = df_filtrado[
                    (df_filtrado[col] >= rango[0]) &
                    (df_filtrado[col] <= rango[1])
                ]

    st.caption(f"{len(df_filtrado)} registros después de filtros")
    st.dataframe(
        df_filtrado[['fecha'] + cols_seleccionadas],
        use_container_width=True, height=300
    )

    csv = df_filtrado[['fecha'] + cols_seleccionadas].to_csv(index=False)
    st.download_button("⬇ Descargar tabla filtrada (CSV)", csv, "datos_filtrados.csv", "text/csv")

    st.markdown("---")
    st.markdown("### Serie temporal")
    col_serie = st.multiselect(
        "Variables a graficar en el tiempo", cols_seleccionadas,
        default=cols_seleccionadas[:2], key="serie_cols"
    )

    if col_serie:
        fig_serie = make_subplots(
            rows=len(col_serie), cols=1, shared_xaxes=True,
            subplot_titles=col_serie, vertical_spacing=0.06
        )
        colores = ['#2d6a9f','#f5820a','#3fb950','#e63946','#8b949e']
        for i, col in enumerate(col_serie):
            fig_serie.add_trace(
                go.Scatter(
                    x=df_filtrado['fecha'], y=df_filtrado[col],
                    mode='lines+markers', name=col,
                    line=dict(color=colores[i % len(colores)], width=1.5),
                    marker=dict(size=5),
                    hovertemplate='%{x|%Y-%m-%d}<br>%{y:.2f}<extra>' + col + '</extra>'
                ), row=i+1, col=1
            )
        fig_serie.update_layout(**layout_claro(height=450))
        fig_serie.update_xaxes(gridcolor='#e0e0e0')
        fig_serie.update_yaxes(gridcolor='#e0e0e0')
        st.plotly_chart(fig_serie, use_container_width=True, key="chart_tab1_serie")

    st.markdown("---")
    st.markdown("### Correlaciones")

    if len(cols_seleccionadas) >= 2:
        col_izq, col_der = st.columns(2)
        with col_izq:
            st.markdown("**Matriz de correlación**")
            corr_matrix = calcular_correlaciones(df_filtrado, cols_seleccionadas)
            fig_heatmap = go.Figure(go.Heatmap(
                z=corr_matrix.values,
                x=corr_matrix.columns.tolist(),
                y=corr_matrix.index.tolist(),
                text=corr_matrix.round(3).values,
                texttemplate='%{text}',
                colorscale=[[0,'#1a3a5c'],[0.5,'#21262d'],[1,'#f5820a']],
                zmin=-1, zmax=1, colorbar=dict(thickness=10)
            ))
            fig_heatmap.update_layout(**layout_claro(height=450))
            st.plotly_chart(fig_heatmap, use_container_width=True, key="chart_tab1_heatmap")

        with col_der:
            st.markdown("**Tabla de correlaciones**")
            corr_df = corr_matrix.stack().reset_index()
            corr_df.columns = ['Variable A', 'Variable B', 'r']
            corr_df = corr_df[corr_df['Variable A'] != corr_df['Variable B']]
            corr_df = corr_df.drop_duplicates(subset=['r']).sort_values('r', ascending=False)
            corr_df['Fuerza'] = corr_df['r'].abs().apply(
                lambda x: '🟢 Fuerte' if x >= 0.7
                else ('🟡 Moderada' if x >= 0.4
                else ('🔴 Débil' if x >= 0.2 else '⚪ Muy débil'))
            )
            st.dataframe(corr_df, use_container_width=True, height=320)

    st.markdown("---")
    st.markdown("### Estadísticas descriptivas")
    stats_df = calcular_stats(df_filtrado, cols_seleccionadas)
    st.dataframe(stats_df, use_container_width=True)

    csv_stats = stats_df.to_csv(index=False)
    st.download_button("⬇ Descargar estadísticas (CSV)", csv_stats, "estadisticas.csv", "text/csv")


# ═════════════════════════════════════════════════════════════════════════
# TAB 2 — AJUSTE AUTOMÁTICO
# ═════════════════════════════════════════════════════════════════════════
with tab2:

    st.markdown("### Configuración del modelo")

    col_cfg1, col_cfg2, col_cfg3 = st.columns(3)

    with col_cfg1:
        y_col = st.selectbox("Variable Y (salida)", cols_numericas, index=0, key="auto_y")

    with col_cfg2:
        modo = st.radio(
            "Modo de entrada", ["Una variable X", "Múltiples variables X"],
            key="auto_modo", horizontal=True
        )

    with col_cfg3:
        grado = st.select_slider(
            "Grado del polinomio", options=[1, 2, 3, 4], value=2,
            key="auto_grado", help="1=lineal, 2=cuadrático, 3=cúbico, 4=cuártico"
        )

    opciones_x = [c for c in cols_numericas if c != y_col]

    if modo == "Una variable X":
        x_cols = [st.selectbox("Variable X (entrada)", opciones_x, key="auto_x_single")]
    else:
        x_cols = st.multiselect(
            "Variables X (entradas)", opciones_x,
            default=opciones_x[:2] if len(opciones_x) >= 2 else opciones_x,
            key="auto_x_multi"
        )
        if x_cols:
            st.caption("ℹ️ Con múltiples variables el polinomio se aplica a cada una por separado")

    st.markdown("---")

    ajustar = st.button("⚙️ Ajustar modelo", type="primary", use_container_width=True)

    if ajustar or 'modelo_resultado' in st.session_state:

        if ajustar:
            if not x_cols:
                st.error("Seleccioná al menos una variable X")
                st.stop()

            df_base   = df_train_global if usar_corte else df_filtrado
            df_modelo = df_base[x_cols + [y_col]].dropna().reset_index(drop=True)

            if len(df_modelo) < 10:
                st.error(
                    f"Muy pocos datos de entrenamiento ({len(df_modelo)} registros). "
                    f"{'Revisá la fecha de corte.' if usar_corte else 'Ampliá el rango de fechas.'}"
                )
                st.stop()

            X_raw    = df_modelo[x_cols].values
            y        = df_modelo[y_col].values
            poly     = PolynomialFeatures(degree=grado, include_bias=False)
            X_poly   = poly.fit_transform(X_raw)
            model    = LinearRegression().fit(X_poly, y)
            y_pred   = model.predict(X_poly)
            r2       = r2_score(y, y_pred)
            residuos = y - y_pred

            st.session_state['modelo_resultado'] = {
                'modo_lab' : False,
                'model'    : model,
                'poly'     : poly,
                'df_modelo': df_modelo,
                'x_cols'   : x_cols,
                'y_col'    : y_col,
                'grado'    : grado,
                'r2'       : r2,
                'y_pred'   : y_pred,
                'residuos' : residuos,
                'modo'     : modo
            }

        res      = st.session_state['modelo_resultado']
        modo_lab = res.get('modo_lab', False)
        df_mod   = res['df_modelo']
        x_cols_r = res['x_cols']
        y_col_r  = res['y_col']
        grado_r  = res['grado']
        r2       = res['r2']
        y_pred   = res['y_pred']
        residuos = res['residuos']
        model    = res['model']
        poly     = res['poly']

        st.markdown("### Resultados del ajuste")
        k1, k2, k3, k4 = st.columns(4)
        k1.metric("R²", f"{r2:.4f}")
        k2.metric("Registros", len(df_mod))
        k3.metric("Grado", grado_r)
        k4.metric("Error std (residuos)", f"{residuos.std():.3f}")

        st.markdown("---")
        st.markdown("### Ecuación del modelo")

        coefs     = model.coef_
        intercept = model.intercept_
        nombres   = poly.get_feature_names_out(x_cols_r)

        def limpiar_nombre(nombre, x_cols):
            res_n = nombre
            for i, col in enumerate(x_cols):
                res_n = res_n.replace(f"x{i}", col)
            return res_n

        eq_parts = []
        for nombre, coef in zip(nombres, coefs):
            nombre_limpio = limpiar_nombre(nombre, x_cols_r)
            signo = "+" if coef >= 0 else "−"
            eq_parts.append(f"{signo} {abs(coef):.6f}·{nombre_limpio}")
        signo_k = "+" if intercept >= 0 else "−"
        eq_str  = f"{y_col_r} = {signo_k} {abs(intercept):.4f} " + " ".join(eq_parts)
        st.code(eq_str, language=None)

        coef_df = pd.DataFrame({
            'Término': ['Término independiente (k)'] + [limpiar_nombre(n, x_cols_r) for n in nombres],
            'Coeficiente': [round(intercept, 6)] + [round(c, 6) for c in coefs]
        })
        st.dataframe(coef_df, use_container_width=True, hide_index=True)

        st.markdown("---")
        st.markdown("### Gráfico")

        if len(x_cols_r) == 1:
            x_col_r = x_cols_r[0]
            fig     = go.Figure()

            if usar_corte and len(df_new_global) > 0:
                df_mod_train = df_train_global[[x_col_r, y_col_r]].dropna()
                df_mod_new   = df_new_global[[x_col_r, y_col_r]].dropna()

                fig.add_trace(go.Scatter(
                    x=df_mod_train[x_col_r], y=df_mod_train[y_col_r], mode='markers',
                    name=f'Train (hasta {fecha_corte})',
                    marker=dict(color='#1a3a5c', size=8, opacity=0.85, line=dict(color='#2d6a9f', width=1)),
                    hovertemplate=f'{x_col_r}: %{{x:.2f}}<br>{y_col_r}: %{{y:.2f}}<extra>Train</extra>'
                ))

                if len(df_mod_new) > 1:
                    fig.add_trace(go.Scatter(
                        x=df_mod_new[x_col_r].iloc[:-1], y=df_mod_new[y_col_r].iloc[:-1],
                        mode='markers', name='Nuevos',
                        marker=dict(color='#f5820a', size=9, opacity=0.85,
                                    line=dict(color='rgba(255,255,255,0.3)', width=0.5)),
                        hovertemplate=f'{x_col_r}: %{{x:.2f}}<br>{y_col_r}: %{{y:.2f}}<extra>Nuevo</extra>'
                    ))

                if len(df_mod_new) >= 1:
                    ultimo = df_mod_new.iloc[-1]
                    fig.add_trace(go.Scatter(
                        x=[ultimo[x_col_r]], y=[ultimo[y_col_r]], mode='markers',
                        name='Último dato',
                        marker=dict(color='#c45200', size=14, symbol='diamond',
                                    line=dict(color='white', width=1.5)),
                        hovertemplate=f'{x_col_r}: %{{x:.2f}}<br>{y_col_r}: %{{y:.2f}}<extra>Último</extra>'
                    ))

                if len(df_mod_new) > 0:
                    x_new_pred   = poly.transform(df_mod_new[[x_col_r]].values)
                    y_new_pred   = model.predict(x_new_pred)
                    desvio_medio = (df_mod_new[y_col_r].values - y_new_pred).mean()
                    st.metric(
                        "Desvío promedio datos nuevos vs modelo",
                        f"{desvio_medio:+.3f}",
                        delta="⚠ Por debajo" if desvio_medio < -0.5 else "✓ Dentro del modelo",
                        delta_color="inverse"
                    )
            else:
                fig.add_trace(go.Scatter(
                    x=df_mod[x_col_r].values, y=df_mod[y_col_r].values,
                    mode='markers', name='Datos',
                    marker=dict(color='#1a3a5c', size=8, opacity=0.8, line=dict(color='#2d6a9f', width=1)),
                    hovertemplate=f'{x_col_r}: %{{x:.2f}}<br>{y_col_r}: %{{y:.2f}}<extra></extra>'
                ))

            x_all   = df_mod[x_col_r].values
            x_curve = np.linspace(x_all.min(), x_all.max(), 300).reshape(-1, 1)
            y_curve = model.predict(poly.transform(x_curve))
            fig.add_trace(go.Scatter(
                x=x_curve.flatten(), y=y_curve, mode='lines',
                name=f'Ajuste grado {grado_r} (R²={r2:.3f})',
                line=dict(color='#e63946', width=2.5)
            ))
            fig.update_layout(**layout_claro(height=450))
            st.plotly_chart(fig, use_container_width=True, key="chart_tab2_single")

        else:
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=df_mod[y_col_r].values, y=y_pred, mode='markers', name='Datos',
                marker=dict(color='#2d6a9f', size=7, opacity=0.8),
                hovertemplate='Real: %{x:.2f}<br>Predicho: %{y:.2f}<extra></extra>'
            ))
            lim = [df_mod[y_col_r].min(), df_mod[y_col_r].max()]
            fig.add_trace(go.Scatter(
                x=lim, y=lim, mode='lines', name='Ideal',
                line=dict(color='#e63946', width=1.5, dash='dash')
            ))
            fig.update_layout(**layout_claro(height=450, titulo=f'Real vs Predicho  —  R² = {r2:.4f}'))
            st.plotly_chart(fig, use_container_width=True, key="chart_tab2_multi")

        with st.expander("📉 Análisis de residuos"):
            fig_res = go.Figure()
            fig_res.add_trace(go.Scatter(
                x=y_pred, y=residuos, mode='markers',
                marker=dict(color='#f5820a', size=7, opacity=0.7),
                hovertemplate='Predicho: %{x:.2f}<br>Residuo: %{y:.2f}<extra></extra>'
            ))
            fig_res.add_hline(y=0, line_dash='dash', line_color='#e63946')
            fig_res.update_layout(**layout_claro(height=300))
            st.plotly_chart(fig_res, use_container_width=True, key="chart_tab2_res")

        st.markdown("---")
        st.markdown("### 🎛️ Simulador")

        sim_cols = st.columns(len(x_cols_r))
        sim_vals = {}
        for i, xc in enumerate(x_cols_r):
            v_min = float(df_mod[xc].min())
            v_max = float(df_mod[xc].max())
            v_med = float(df_mod[xc].mean())
            with sim_cols[i]:
                sim_vals[xc] = st.slider(
                    xc, min_value=round(v_min, 2), max_value=round(v_max, 2),
                    value=round(v_med, 2), step=round((v_max - v_min) / 100, 2),
                    key=f"sim_{xc}"
                )

        X_sim      = np.array([[sim_vals[xc] for xc in x_cols_r]])
        X_sim_poly = poly.transform(X_sim)
        y_sim      = model.predict(X_sim_poly)[0]
        y_mean     = df_mod[y_col_r].mean()
        delta      = y_sim - y_mean

        s1, s2, s3 = st.columns(3)
        s1.metric(f"{y_col_r} predicho", f"{y_sim:.3f}", delta=f"{delta:+.3f} vs promedio")
        s2.metric("Promedio histórico", f"{y_mean:.3f}")
        s3.metric("R² del modelo", f"{r2:.4f}")

        st.markdown("---")
        st.markdown("### 📥 Exportar modelo")

        col_exp1, col_exp2 = st.columns(2)

        with col_exp1:
            if st.button("📊 Generar Excel de coeficientes", use_container_width=True, key="btn_excel_tab2"):
                buffer = exportar_coeficientes_excel(st.session_state['modelo_resultado'], desde, hasta)
                st.download_button(
                    label="⬇ Descargar Excel", data=buffer,
                    file_name=f"modelo_{y_col_r}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, key="dl_excel_tab2"
                )

        with col_exp2:
            if st.button("📄 Generar reporte HTML", use_container_width=True, key="btn_html_tab2"):
                html = generar_reporte_html(
                    df_filtrado, st.session_state['modelo_resultado'],
                    cols_seleccionadas, titulo="Reporte Metalúrgico — Ajuste Automático"
                )
                st.download_button(
                    label="⬇ Descargar HTML", data=html,
                    file_name=f"reporte_met_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.html",
                    mime="text/html", use_container_width=True, key="dl_html_tab2"
                )


# ═════════════════════════════════════════════════════════════════════════
# TAB 3 — MODELO MANUAL
# ═════════════════════════════════════════════════════════════════════════
with tab3:

    st.markdown("### Configuración")

    usar_ec_lab_tab3 = st.toggle(
        "⚗ Usar ecuación de laboratorio (lgo/mchch)", value=False,
        key="switch_lab_tab3", help="Precarga los coeficientes con los valores de laboratorio"
    )

    if usar_ec_lab_tab3:
        st.info("Coeficientes precargados con valores de laboratorio. Podés modificarlos libremente.")

    col_m1, col_m2 = st.columns(2)

    with col_m1:
        if usar_ec_lab_tab3:
            cols_lgo_t3 = [
                c for c in cols_numericas
                if any(k in c.lower() for k in ['lgo', 'pct_lgo', '%lgo', 'porc_lgo'])
            ]
            x_col_man = st.selectbox(
                "Columna % lgo", cols_numericas,
                index=cols_numericas.index(cols_lgo_t3[0]) if cols_lgo_t3 else 0,
                key="man_x_lab"
            )
        else:
            x_col_man = st.selectbox("Variable X", cols_numericas, key="man_x")

    with col_m2:
        y_col_man = st.selectbox(
            "Variable Y", [c for c in cols_numericas if c != x_col_man], key="man_y"
        )

    if usar_ec_lab_tab3:
        grado_man = 3
        st.caption("Grado fijo en 3 para ecuación de laboratorio")
    else:
        grado_man = st.select_slider(
            "Grado del polinomio", options=[1, 2, 3, 4], value=2, key="man_grado"
        )

    st.markdown("---")
    st.markdown("### Coeficientes")

    if usar_ec_lab_tab3:
        coefs_pre = ECUACIONES_LAB.get(y_col_man, {"k": 0.0, "c1": 0.0, "c2": 0.0, "c3": 0.0})
        defaults  = [coefs_pre["k"], coefs_pre["c1"], coefs_pre["c2"], coefs_pre["c3"]]
        st.caption("⚗ Coeficientes precargados desde ecuación de laboratorio")

        _clave_carga_t3 = f"_lab_cargado_t3_{y_col_man}"
        if _clave_carga_t3 not in st.session_state:
            for key in [f"coef_{i}_tab3" for i in range(4)]:
                if key in st.session_state:
                    del st.session_state[key]
            st.session_state["coef_0_tab3"] = coefs_pre["k"]
            st.session_state["coef_1_tab3"] = coefs_pre["c1"]
            st.session_state["coef_2_tab3"] = coefs_pre["c2"]
            st.session_state["coef_3_tab3"] = coefs_pre["c3"]
            st.session_state[_clave_carga_t3] = True
    else:
        defaults = [0.0] * (grado_man + 1)
        if 'modelo_resultado' in st.session_state:
            res_auto = st.session_state['modelo_resultado']
            if (not res_auto.get('modo_lab', False) and
                res_auto['grado'] == grado_man and
                len(res_auto['x_cols']) == 1 and
                res_auto['x_cols'][0] == x_col_man and
                res_auto['y_col'] == y_col_man):
                defaults = (
                    [float(res_auto['model'].intercept_)] +
                    [float(c) for c in res_auto['model'].coef_]
                )

    nombres_terminos = {
        0: 'k — Término independiente',
        1: 'c₁ · x', 2: 'c₂ · x²', 3: 'c₃ · x³', 4: 'c₄ · x⁴',
    }

    coef_manuales = {}
    cols_coef     = st.columns(grado_man + 1)
    for i in range(grado_man + 1):
        with cols_coef[i]:
            coef_manuales[i] = st.number_input(
                nombres_terminos[i],
                value=defaults[i] if i < len(defaults) else 0.0,
                format="%.6f", key=f"coef_{i}_tab3"
            )

    st.markdown("---")
    st.markdown("### Vista previa de la ecuación")

    partes_eq = []
    for i in range(grado_man, -1, -1):
        c = coef_manuales[i]
        if i == 0:
            partes_eq.append(f"{c:+.4f}")
        elif i == 1:
            partes_eq.append(f"{c:+.6f}·{x_col_man}")
        else:
            partes_eq.append(f"{c:+.6f}·{x_col_man}^{i}")
    eq_preview = f"{y_col_man} = " + " ".join(partes_eq)
    st.code(eq_preview, language=None)

    def predict_manual(x, coefs, grado):
        return sum(coefs[i] * x**i for i in range(grado + 1))

    st.markdown("---")
    st.markdown("### Gráfico")

    df_man    = df_filtrado[[x_col_man, y_col_man]].dropna()
    color_txt = '#1a1a2e' if FONDO_CLARO else '#e6edf3'

    if len(df_man) > 0:
        x_vals = df_man[x_col_man].values
        y_vals = df_man[y_col_man].values

        x_curve    = np.linspace(x_vals.min(), x_vals.max(), 300)
        y_curve    = np.array([predict_manual(x, coef_manuales, grado_man) for x in x_curve])
        y_pred_man = np.array([predict_manual(x, coef_manuales, grado_man) for x in x_vals])
        r2_man     = r2_score(y_vals, y_pred_man)

        fig_man = go.Figure()
        fig_man.add_trace(go.Scatter(
            x=x_vals, y=y_vals, mode='markers', name='Datos reales',
            marker=dict(color='#1a3a5c', size=8, opacity=0.8, line=dict(color='#2d6a9f', width=1)),
            hovertemplate=f'{x_col_man}: %{{x:.2f}}<br>{y_col_man}: %{{y:.2f}}<extra></extra>'
        ))

        nombre_curva = "Ec. laboratorio" if usar_ec_lab_tab3 else "Modelo manual"
        fig_man.add_trace(go.Scatter(
            x=x_curve, y=y_curve, mode='lines',
            name=f'{nombre_curva} (R²={r2_man:.3f})',
            line=dict(color='#f5820a', width=2.5)
        ))

        if 'modelo_resultado' in st.session_state:
            res_auto = st.session_state['modelo_resultado']
            if (not res_auto.get('modo_lab', False) and
                len(res_auto['x_cols']) == 1 and
                res_auto['x_cols'][0] == x_col_man and
                res_auto['y_col'] == y_col_man):
                poly_auto    = res_auto['poly']
                model_auto   = res_auto['model']
                y_curve_auto = model_auto.predict(poly_auto.transform(x_curve.reshape(-1, 1)))
                fig_man.add_trace(go.Scatter(
                    x=x_curve, y=y_curve_auto, mode='lines',
                    name=f'Ajuste automático (R²={res_auto["r2"]:.3f})',
                    line=dict(color='#3fb950', width=1.5, dash='dash')
                ))

        fig_man.update_layout(**layout_claro(height=450))
        st.plotly_chart(fig_man, use_container_width=True, key="chart_tab3_manual")

        st.markdown("---")
        st.markdown("### 🎛️ Simulador")

        v_min = float(x_vals.min())
        v_max = float(x_vals.max())
        v_med = float(x_vals.mean())

        sim_x_man = st.slider(
            x_col_man, min_value=round(v_min, 2), max_value=round(v_max, 2),
            value=round(v_med, 2), step=round((v_max - v_min) / 100, 2),
            key="sim_manual_x"
        )
        sim_x_input = st.number_input(
            "O ingresá un valor exacto", value=float(sim_x_man),
            format="%.2f", key="sim_manual_input"
        )

        x_sim_final = sim_x_input
        y_sim_man   = predict_manual(x_sim_final, coef_manuales, grado_man)
        y_mean_man  = float(y_vals.mean())
        delta_man   = y_sim_man - y_mean_man

        s1, s2, s3 = st.columns(3)
        s1.metric(f"{y_col_man} predicho", f"{y_sim_man:.3f}", delta=f"{delta_man:+.3f} vs promedio")
        s2.metric("Promedio histórico", f"{y_mean_man:.3f}")
        s3.metric("R² modelo manual", f"{r2_man:.4f}")

        st.markdown("---")
        st.markdown("### 📥 Exportar modelo")

        if st.button("📄 Generar reporte HTML", use_container_width=True, key="btn_html_tab3"):
            poly_man         = PolynomialFeatures(degree=grado_man, include_bias=False)
            poly_man.fit_transform(df_man[[x_col_man]].values)
            m_man            = LinearRegression()
            m_man.coef_      = np.array([coef_manuales[i] for i in range(1, grado_man + 1)])
            m_man.intercept_ = coef_manuales[0]

            resultado_man = {
                'modo_lab' : usar_ec_lab_tab3,
                'model'    : m_man,
                'poly'     : poly_man,
                'df_modelo': df_man,
                'x_cols'   : [x_col_man],
                'y_col'    : y_col_man,
                'grado'    : grado_man,
                'r2'       : r2_man,
                'y_pred'   : y_pred_man,
                'residuos' : y_vals - y_pred_man,
                'modo'     : 'manual'
            }
            html = generar_reporte_html(
                df_filtrado, resultado_man, cols_seleccionadas,
                titulo="Reporte Metalúrgico — Modelo Manual"
            )
            st.download_button(
                label="⬇ Descargar HTML", data=html,
                file_name=f"reporte_met_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.html",
                mime="text/html", use_container_width=True, key="dl_html_tab3"
            )

    else:
        st.warning("No hay datos suficientes para las variables seleccionadas.")